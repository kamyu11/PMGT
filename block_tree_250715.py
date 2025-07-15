import streamlit as st
import pandas as pd
import re
import os
import tempfile
from ctypes import CDLL, WinDLL, c_int
from openpyxl import load_workbook
import xlrd

# ───────────────────────────────────────────────────
# ▶ DRM 해제 함수 정의
#    - Fasoo DRM용 DLL을 로드하여 암호화 해제(EnableDRM) 수행
#    - 여러 후보 함수명을 순차적으로 시도하고, 실패 시 Streamlit 경고창 표시
#    - 반환값 ret == 0 이면 해제 실패로 간주
# ───────────────────────────────────────────────────
def setting_fasoo():
    dll_path = r"C:\Windows\System32\f_nxldr.dll"  # 시스템에 설치된 DRM DLL 경로
    try:
        # 1) CDLL/CDECL 호출 방식 시도 → 실패 시 WinDLL(STDCALL) 방식
        try:
            fasoo = CDLL(dll_path)
        except OSError:
            fasoo = WinDLL(dll_path)

        # 2) 함수 이름 후보 순회 (일반, 오타, 장식된 이름)
        for func_name in ("EnableDRM", "EnablDRM", "EnablDRM@0"):
            try:
                # 해당 이름의 함수 심볼 조회
                func = getattr(fasoo, func_name)
                func.restype = c_int  # 반환값을 int로 해석
                ret = func()          # DLL 함수 호출
                break
            except AttributeError:
                # 다음 후보 함수명으로 계속 시도
                continue
        else:
            # 모든 후보 실패 시 경고 후 False 반환
            st.warning(f"DRM 함수가 '{dll_path}' DLL 내에 없습니다.")
            return False

        # 3) 반환값 확인: 0은 실패, 그 외는 성공
        if ret == 0:
            st.warning("암호화 해제에 실패했습니다 (ret==0).")
            return False

    except FileNotFoundError:
        # DLL 파일 자체를 찾을 수 없을 때
        st.warning("DRM DLL 파일이 없어 DRM 기능이 비활성화됩니다.")
        return False
    except Exception as e:
        # 기타 예외 상황 처리
        st.warning(f"DRM 설정 중 오류 발생: {e}")
        return False

    # 모든 단계 통과 시 True 반환
    return True

# ───────────────────────────────────────────────────
# ▶ 0) 초기 DRM 해제 시도
#    - 실패하더라도 이후 로직은 계속 실행되나, 암호화된 데이터는 처리 불가할 수 있음
# ───────────────────────────────────────────────────
if not setting_fasoo():
    # 여기에 DRM 실패 후 추가 로직 삽입 가능
    pass

# ───────────────────────────────────────────────────
# ▶ 1) Streamlit 앱 제목 표시
# ───────────────────────────────────────────────────
st.title('블록 계층 구조 데이터화 도구')

# ───────────────────────────────────────────────────
# ▶ 2) 사용자 입력 위젯: 파일 업로드 및 결과 파일명 입력
#    - 기존 Tkinter 다이얼로그 대신 Streamlit의 file_uploader 사용
#    - 입력된 파일명을 기반으로 임시 파일 생성 후 처리
# ───────────────────────────────────────────────────
# 2-1) Excel 파일 업로드 (.xlsx, .xls)
uploaded = st.file_uploader(
    label='1) Block Division Excel 파일 업로드',
    type=['xlsx', 'xls']
)
# 2-2) 결과 저장용 파일명 텍스트 입력
save_name = st.text_input(
    label='2) 결과 저장 파일명 입력',
    value='SN0000_Block Hierachy_result.xlsx'
)

# ───────────────────────────────────────────────────
# ▶ 3) 실행 버튼 눌렀을 때 처리 흐름
# ───────────────────────────────────────────────────
if uploaded and save_name and st.button('실행'):
    # 3-1) 업로드된 바이너리를 임시 파일에 쓰기
    suffix = os.path.splitext(uploaded.name)[1]  # .xlsx 또는 .xls
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(uploaded.getvalue())
        file_path = tmp.name

    # 3-2) 저장 경로 설정 (현재 작업 디렉토리)
    save_path = os.path.join(os.getcwd(), save_name)

    # ───────────────────────────────────────────────────
    # ▶ 4) pandas + openpyxl/xlrd 로 원본 시트 로드 및 병합 범위 수집
    # ───────────────────────────────────────────────────
    if file_path.lower().endswith('.xlsx'):
        # 4-1) .xlsx 파일 처리
        df_full = pd.read_excel(
            file_path,
            sheet_name=1,
            header=None,
            engine='openpyxl'
        )
        wb_tmp = load_workbook(file_path, data_only=True)
        ws_tmp = wb_tmp.worksheets[1] if len(wb_tmp.worksheets) > 1 else wb_tmp.worksheets[0]
        merged_ranges = [
            (r.min_row, r.min_col, r.max_row, r.max_col)
            for r in ws_tmp.merged_cells.ranges
        ]
        wb_tmp.close()
    elif file_path.lower().endswith('.xls'):
        # 4-2) .xls 파일 처리
        try:
            book_temp = xlrd.open_workbook(file_path, formatting_info=True)
            sheet_temp = book_temp.sheet_by_index(1)
        except Exception as e:
            st.error(f".xls 파일 로드 오류: {e}")
            st.stop()
        df_full = pd.read_excel(
            file_path,
            sheet_name=1,
            header=None,
            engine='xlrd'
        )
        merged_ranges = [
            (rlo+1, clo+1, rhi, chi)
            for (rlo, rhi, clo, chi) in sheet_temp.merged_cells
        ]
    else:
        st.error('지원되지 않는 파일 형식입니다.')
        st.stop()

    # ───────────────────────────────────────────────────
    # ▶ 5) "원본 DataFrame"에서 5행(1-based) 이후만 사용
    #    - 0-based index로는 START_ROW_IDX=4 (5번째 행)
    # ───────────────────────────────────────────────────
    START_ROW_IDX = 4
    df_data = df_full.iloc[START_ROW_IDX:].reset_index(drop=True)
    df_data.columns = [f'COLUMN{i+1}' for i in range(df_data.shape[1])]

    # ───────────────────────────────────────────────────
    # ▶ 6) 병합된 셀 영역만 "위쪽 값으로 채우기"
    #    - merged_ranges 목록의 각 영역을 순회하며 빈칸(NaN)만 상단값 복사
    # ───────────────────────────────────────────────────
    def fill_merged(df, merges, start_idx):
        for min_row, min_col, max_row, max_col in merges:
            sr = min_row - 1 - start_idx
            sc = min_col - 1
            er = max_row - 1 - start_idx
            ec = max_col - 1
            # 원본 df_full에서 최상단 값 가져오기
            if 0 <= min_row-1 < len(df_full) and 0 <= min_col-1 < df_full.shape[1]:
                top_val = df_full.iat[min_row-1, min_col-1]
            else:
                continue
            # df_data 범위 내 빈칸에만 채우기
            for r in range(sr, er+1):
                for c in range(sc, ec+1):
                    if 0 <= r < len(df) and 0 <= c < df.shape[1] and pd.isna(df.iat[r, c]):
                        df.iat[r, c] = top_val
    fill_merged(df_data, merged_ranges, START_ROW_IDX)

    # ───────────────────────────────────────────────────
    # ▶ 7) 부모-자식 관계 추출
    #    - G열 (COLUMN7): 접미사 P/C/S 판단하여 추가
    #    - F~B 열: 각 열을 자식으로, 왼쪽 첫 비어있지 않은 열을 부모로
    # ───────────────────────────────────────────────────
    rows = []
    def clean_name(val):
        if pd.isna(val): return None
        return re.sub(r'[()\s\n]', '', str(val))[:5]
    parent_cols = ['COLUMN6','COLUMN5','COLUMN4','COLUMN3','COLUMN2','COLUMN1']
    # 7-1) G열 처리
    for _, row in df_data.iterrows():
        raw = row.get('COLUMN7')
        if pd.isna(raw): continue
        parent = next((clean_name(row[pc]) for pc in parent_cols if pd.notna(row.get(pc))), None)
        if not parent: continue
        base = re.sub(r'[\n\s]', '', str(raw)).strip()
        suffixes = [s for idx, s in enumerate(['P','C','S'], start=8)
                    if df_data.shape[1]>=idx and pd.notna(pd.to_numeric(row.get(f'COLUMN{idx}'), errors='coerce'))]
        for s in suffixes or ['']:
            rows.append({'부모': parent, '자식': f"{base}{s}"})
    # 7-2) F~B열 처리
    for i in range(6, 1, -1):
        child_col = f'COLUMN{i}'
        pr_cols = [f'COLUMN{j}' for j in range(i-1, 0, -1)]
        for _, row in df_data.iterrows():
            raw = row.get(child_col)
            if pd.isna(raw): continue
            child = clean_name(raw)
            parent = next((clean_name(row[p]) for p in pr_cols if pd.notna(row.get(p))), None)
            if parent:
                rows.append({'부모': parent, '자식': child})

    result_df = pd.DataFrame(rows).drop_duplicates().reset_index(drop=True)

    # ── 8) Remark 분리 (줄바꿈 인식) ──
    remark_rows = []
    if 'COLUMN11' in df_data.columns:
        for raw in df_data['COLUMN11'].dropna().astype(str):
            for line in raw.splitlines():
                line = line.strip()
                if not line:
                    continue
                parts = line.split(':', 1)
                parts = line.split(':', 1)
                # 괄호 문자 제거
                block = re.sub(r'[()]', '', parts[0].strip())
                remark = parts[1].strip() if len(parts) > 1 else ''
                remark_rows.append({'Block명': block, 'Remark': remark})
    remark_df = pd.DataFrame(remark_rows) \
        .drop_duplicates() \
        .reset_index(drop=True)

    out = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    with pd.ExcelWriter(out.name, engine='openpyxl') as writer:
        result_df.to_excel(writer, sheet_name='Result', index=False)
        if not remark_df.empty:
            remark_df.to_excel(writer, sheet_name='Remark', index=False)

    # 9) 결과 저장 (두 개 시트)
    out = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    with pd.ExcelWriter(out.name, engine='openpyxl') as writer:
        result_df.to_excel(writer, sheet_name='Result', index=False)
        if not remark_df.empty:
            remark_df.to_excel(writer, sheet_name='Remark', index=False)

    # 10) 다운로드 버튼
    st.download_button(
        label='결과 다운로드',
        data=open(out.name, 'rb').read(),
        file_name=save_name,
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    st.success('엑셀 파일 저장이 완료되었습니다!')
